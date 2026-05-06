from __future__ import annotations

import copy
import json
import os
import re
import subprocess
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from importlib.util import find_spec
from pathlib import Path
from typing import Any

from ..artifact_store import StageWorkspace
from ..config import project_root
from ..test_environment import load_test_environment
from .allure_helper import build_allure_viewing_guide, generate_allure_html_report, parse_pytest_summary
from .base_agent import BaseAgent
from .test_generator import (
    build_generated_test_file,
    deep_merge_payload,
    derive_base_url,
    extract_request_schema,
    schema_blank_payload,
)


class AutomationAgent(BaseAgent):
    agent_name = "automation_agent"
    stage_name = "automation_execution"
    _CONTACT_WAY_DEFAULT_CATEGORY_ID = 16
    _CONTACT_WAY_DEFAULT_USER_IDS = ("74698", "13951756117")
    _CONTACT_WAY_DEFAULT_BACKUP_USER_ID = "13229"

    def __init__(self, test_environment_config_path: Path | None = None) -> None:
        self.test_environment_config_path = test_environment_config_path

    def run(
        self,
        run_id: str,
        run_dir: Path,
        endpoint_mapping: dict[str, Any],
        stage_workspace: StageWorkspace,
    ) -> dict[str, object]:
        template_config = self._load_agent4_template_config()
        generated_tests, data_resolution = self._prepare_generated_tests(
            endpoint_mapping["mappings"],
            run_id=run_id,
            template_config=template_config,
        )
        test_file = stage_workspace.write_output_text(
            "generated_tests/test_generated_api_cases.py",
            build_generated_test_file(generated_tests),
        )

        automation_plan = {
            "run_id": run_id,
            "mode": "execute_with_skip_when_api_base_url_missing",
            "generated_tests": generated_tests,
            "template_config_summary": self._redact_template_config(template_config),
            "data_resolution": data_resolution,
        }
        stage_workspace.write_output_json("automation_plan.json", automation_plan)

        allure_results_dir = stage_workspace.output_dir / "allure-results"
        pytest_command = [sys.executable, "-m", "pytest", str(test_file.resolve()), "-q"]
        if find_spec("allure_pytest") is not None:
            pytest_command.append(f"--alluredir={allure_results_dir}")

        completed = subprocess.run(
            pytest_command,
            cwd=run_dir,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            check=False,
            env=self._build_pytest_env(template_config=template_config),
        )
        summary = parse_pytest_summary(completed.stdout + completed.stderr)
        html_report = generate_allure_html_report(allure_results_dir, stage_workspace.output_dir / "allure-report")
        viewing_guide_path = None
        if html_report.get("index_path"):
            viewing_guide = build_allure_viewing_guide(
                report_dir=Path(str(html_report["report_dir"])),
                results_dir=allure_results_dir,
            )
            viewing_guide_path = stage_workspace.write_output_text(
                "allure_report_viewing_guide.md",
                viewing_guide,
            )

        execution_report = {
            "run_id": run_id,
            "pytest_exit_code": completed.returncode,
            "generated_test_count": len(generated_tests),
            "summary": summary,
            "allure_results_dir": str(allure_results_dir),
            "allure_html_report": html_report,
            "allure_viewing_guide": str(viewing_guide_path) if viewing_guide_path else None,
            "stdout": completed.stdout,
            "stderr": completed.stderr,
        }
        stage_workspace.write_output_json("execution_report.json", execution_report)

        failure_summary = "\n".join(
            [
                "# Failure Summary",
                "",
                f"- pytest exit code: {completed.returncode}",
                f"- passed: {summary['passed']}",
                f"- failed: {summary['failed']}",
                f"- skipped: {summary['skipped']}",
            ]
        )
        stage_workspace.write_output_text("failure_summary.md", failure_summary)
        stage_workspace.write_log(
            "execution.log",
            "\n".join(
                [
                    f"run_id={run_id}",
                    f"pytest_exit_code={completed.returncode}",
                    "--- stdout ---",
                    completed.stdout,
                    "--- stderr ---",
                    completed.stderr,
                ]
            ),
        )
        return {
            "automation_plan": automation_plan,
            "execution_report": execution_report,
            "generated_test_file": str(test_file),
        }

    def _default_agent4_template_path(self) -> Path:
        return project_root() / "workspace" / "templates" / "Agent4自动化执行输入模板_鉴权约束修正版.xlsx"

    def _load_agent4_template_config(self) -> dict[str, str]:
        template_path = self._default_agent4_template_path()
        if not template_path.exists():
            return {}
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {}

        workbook = load_workbook(template_path, data_only=True)
        field_map: dict[str, str] = {}
        for sheet_name in ["最小输入", "登录鉴权配置", "可选接口补充"]:
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            for row in range(5, sheet.max_row + 1):
                field_name = sheet.cell(row, 2).value
                field_value = sheet.cell(row, 4).value
                if field_name in {None, ""}:
                    continue
                normalized_name = str(field_name).strip()
                normalized_value = "" if field_value is None else str(field_value).strip()
                field_map[normalized_name] = normalized_value

        config: dict[str, str] = {}
        mappings = {
            "测试环境地址": "API_BASE_URL",
            "登录账号": "FRONT_USERNAME",
            "登录密码": "FRONT_PASSWORD",
            "系统类型 sysType": "FRONT_SYS_TYPE",
            "登录接口地址": "FRONT_LOGIN_PATH",
            "登录接口请求方式": "FRONT_LOGIN_METHOD",
            "登录接口请求头": "FRONT_LOGIN_HEADERS_JSON",
            "登录接口请求体": "FRONT_LOGIN_BODY_TEMPLATE",
            "登录成功标识": "LOGIN_SUCCESS_MARKER",
            "鉴权值提取字段路径": "FRONT_TOKEN_JSON_PATH",
            "鉴权 Header 名称": "AUTH_HEADER_NAME",
            "鉴权 Header 前缀": "AUTH_SCHEME",
        }
        for field_name, env_name in mappings.items():
            value = field_map.get(field_name, "")
            if value:
                config[env_name] = value

        success_marker = config.get("LOGIN_SUCCESS_MARKER", "")
        if success_marker.startswith("code="):
            config["SUCCESS_RESPONSE_CODE"] = success_marker.split("=", 1)[1].strip()
        elif success_marker:
            config["SUCCESS_RESPONSE_CODE"] = success_marker
        return config

    def _redact_template_config(self, template_config: dict[str, str]) -> dict[str, str]:
        redacted: dict[str, str] = {}
        for key, value in template_config.items():
            if key in {"FRONT_PASSWORD"}:
                redacted[key] = "***"
            else:
                redacted[key] = value
        return redacted

    def _prepare_generated_tests(
        self,
        mappings: list[dict[str, Any]],
        *,
        run_id: str,
        template_config: dict[str, str],
    ) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        prepared_cases = copy.deepcopy(mappings)
        report: dict[str, Any] = {"warnings": [], "case_overrides": []}
        if not any(str(case.get("test_case_id", "")).startswith("TC-LIVECODE-") for case in prepared_cases):
            return prepared_cases, report

        try:
            context = self._discover_livecode_context(template_config)
        except Exception as exc:  # pragma: no cover - network/runtime dependent
            report["warnings"].append(f"livecode_context_unavailable: {exc}")
            return prepared_cases, report

        sequence = 0
        for case in prepared_cases:
            test_case_id = str(case.get("test_case_id", ""))
            path = str(case.get("path", ""))
            if str(case.get("module", "")).startswith("微信服务（/wx/mp/）"):
                case["base_url"] = context["wx_base_url"]

            if path == "/contact/way/save":
                sequence += 1
                payload = self._build_contact_way_payload(
                    run_id=run_id,
                    sequence=sequence,
                    original_payload=case.get("payload", {}),
                    request_schema=extract_request_schema(case.get("request_body", {})),
                    context=context,
                )
                if test_case_id == "TC-LIVECODE-002":
                    fixture_id = self._create_contact_way_fixture(payload, context)
                    if fixture_id is not None:
                        payload["id"] = fixture_id
                case["payload"] = payload
                case["payload_strategy"] = "raw_payload"
                report["case_overrides"].append({"test_case_id": test_case_id, "strategy": "contact_way_live_payload"})
                continue

            if path == "/acquisition/link/saveOrUpdate":
                sequence += 1
                case["payload"] = self._build_acquisition_link_payload(
                    run_id=run_id,
                    sequence=sequence,
                    original_payload=case.get("payload", {}),
                    request_schema=extract_request_schema(case.get("request_body", {})),
                    context=context,
                )
                report["case_overrides"].append(
                    {"test_case_id": test_case_id, "strategy": "acquisition_link_live_payload"}
                )
                continue

            if path == "/acquisition/user/online/change":
                online_candidates = context.get("online_candidates", [])
                if not online_candidates:
                    case["skip_reason"] = "未查询到可操作的企微员工，跳过上下线接口自动化。"
                    report["case_overrides"].append(
                        {"test_case_id": test_case_id, "strategy": "skip_no_online_candidates"}
                    )
                    continue
                selected = online_candidates[0]
                payload = dict(case.get("payload", {})) if isinstance(case.get("payload", {}), dict) else {}
                payload["corpId"] = selected["corpId"]
                payload["userId"] = selected["userId"]
                case["payload"] = payload
                report["case_overrides"].append({"test_case_id": test_case_id, "strategy": "online_change_payload"})

        return prepared_cases, report

    def _discover_livecode_context(self, template_config: dict[str, str]) -> dict[str, Any]:
        env = self._build_pytest_env(template_config=template_config)
        operation_base_url = str(env.get("API_BASE_URL", "")).strip()
        if not operation_base_url:
            raise RuntimeError("missing API_BASE_URL")
        wx_base_url = derive_base_url(operation_base_url, "/wx/mp/")
        auth_headers = self._runtime_login_headers(env)
        corp_list = self._request_json("GET", f"{wx_base_url}wechatinfo/corp", headers=auth_headers)
        corp_data = corp_list.get("data", []) if isinstance(corp_list.get("data"), list) else []
        if not corp_data:
            raise RuntimeError("no_available_corps")
        corp = corp_data[0]
        corp_id = str(corp.get("appId") or "")
        if not corp_id:
            raise RuntimeError("corp_id_missing")

        user_list_payload = {"corpId": corp_id, "pageNum": 1, "pageSize": 200}
        user_list = self._request_json("POST", f"{wx_base_url}cp/user/list", payload=user_list_payload, headers=auth_headers)
        users = user_list.get("data", {}).get("list", []) if isinstance(user_list.get("data"), dict) else []
        if not users:
            raise RuntimeError("no_available_cp_users")
        selected_user = self._select_preferred_user(users)

        contact_category_response = self._request_json(
            "GET",
            f"{operation_base_url.rstrip('/')}/contact/category/queryCategoryList?{urllib.parse.urlencode({'corpId': corp_id})}",
            headers=auth_headers,
        )
        contact_categories = self._flatten_tree_nodes(contact_category_response.get("data", []))
        if not contact_categories:
            raise RuntimeError("no_contact_categories")

        link_category_response = self._request_json(
            "GET",
            f"{operation_base_url.rstrip('/')}/acquisition/link/category/selectList?{urllib.parse.urlencode({'corpId': corp_id})}",
            headers=auth_headers,
        )
        link_categories = self._flatten_tree_nodes(link_category_response.get("data", []))
        online_response = self._request_json(
            "GET",
            f"{operation_base_url.rstrip('/')}/acquisition/user/online/list",
            headers=auth_headers,
        )
        online_candidates = online_response.get("data", []) if isinstance(online_response.get("data"), list) else []

        return {
            "operation_base_url": operation_base_url if operation_base_url.endswith("/") else f"{operation_base_url}/",
            "wx_base_url": wx_base_url,
            "auth_headers": auth_headers,
            "corp": corp,
            "corp_id": corp_id,
            "users": users,
            "selected_user": selected_user,
            "contact_categories": contact_categories,
            "link_categories": link_categories,
            "online_candidates": online_candidates,
        }

    def _build_contact_way_payload(
        self,
        *,
        run_id: str,
        sequence: int,
        original_payload: object,
        request_schema: dict[str, Any] | None,
        context: dict[str, Any],
    ) -> dict[str, Any]:
        payload = dict(original_payload) if isinstance(original_payload, dict) else {}
        unique_name = f"auto_{run_id}_{sequence}_{int(time.time())}"[-30:]
        primary_users = self._select_contact_way_primary_users(context["users"])
        backup_user = self._select_contact_way_backup_user(context["users"], primary_users)
        base_payload = {
            "id": payload.get("id"),
            "name": payload.get("name") or unique_name,
            "categoryId": payload.get("categoryId", self._CONTACT_WAY_DEFAULT_CATEGORY_ID),
            "corpId": payload.get("corpId") or context["corp_id"],
            "welcomeConfigVO": payload.get("welcomeConfigVO")
            or {
                "id": None,
                "configType": 3,
            },
            "markTagIdList": payload.get("markTagIdList", []),
            "effectiveType": payload.get("effectiveType", 0),
            "shuntType": payload.get("shuntType", 2),
            "skipVerify": payload.get("skipVerify", 0),
            "nextAutoEnableTime": payload.get("nextAutoEnableTime"),
            "allTimeUsers": payload.get("allTimeUsers")
            or [
                {
                    "userId": user["userId"],
                    "enable": 1,
                    "addNumber": -1,
                    "sort": index,
                }
                for index, user in enumerate(primary_users)
            ],
            "splitTimeUsers": payload.get("splitTimeUsers")
            or [
                {
                    "id": None,
                    "contants": [],
                    "CCCC": [],
                    "DDDD": [],
                }
            ],
            "backupUserId": payload.get("backupUserId") or backup_user["userId"],
        }
        return deep_merge_payload(base_payload, payload)

    def _build_acquisition_link_payload(
        self,
        *,
        run_id: str,
        sequence: int,
        original_payload: object,
        request_schema: dict[str, Any] | None,
        context: dict[str, Any],
    ) -> dict[str, Any]:
        payload = dict(original_payload) if isinstance(original_payload, dict) else {}
        schema_payload = schema_blank_payload(request_schema or {})
        unique_name = f"auto_link_{run_id}_{sequence}_{int(time.time())}"[-30:]
        selected_user = context["selected_user"]
        selected_category = context["link_categories"][0]
        schema_payload.update(
            {
                "createTime": None,
                "updateTime": None,
                "id": None,
                "name": payload.get("name") or unique_name,
                "corpId": payload.get("corpId") or context["corp_id"],
                "categoryId": payload.get("categoryId") or selected_category["value"],
                "markTagIdList": [],
                "skipVerify": payload.get("skipVerify", 1),
                "effectiveType": payload.get("effectiveType", 0),
                "shuntType": payload.get("shuntType", 0),
                "allTimeUsers": payload.get("allTimeUsers")
                or [
                    {
                        "id": None,
                        "userId": selected_user["userId"],
                        "addNumber": -1,
                        "enable": 1,
                        "sort": None,
                        "pointer": None,
                    }
                ],
                "splitTimeUsers": [],
                "backupUserId": None,
                "welcomeConfigVO": None,
                "nextAutoEnableTime": payload.get("nextAutoEnableTime"),
                "isShowOnPage": payload.get("isShowOnPage", 0),
            }
        )
        return deep_merge_payload(schema_payload, payload)

    def _create_contact_way_fixture(self, payload: dict[str, Any], context: dict[str, Any]) -> int | None:
        response = self._request_json(
            "POST",
            f"{context['operation_base_url']}contact/way/save",
            payload=payload,
            headers=context["auth_headers"],
        )
        data = response.get("data")
        return int(data) if isinstance(data, int) else None

    def _flatten_tree_nodes(self, nodes: object) -> list[dict[str, Any]]:
        flattened: list[dict[str, Any]] = []
        if not isinstance(nodes, list):
            return flattened
        stack = [item for item in nodes if isinstance(item, dict)]
        while stack:
            node = stack.pop(0)
            flattened.append(node)
            children = node.get("children", [])
            if isinstance(children, list):
                stack.extend(child for child in children if isinstance(child, dict))
        return flattened

    def _select_preferred_user(self, users: list[dict[str, Any]]) -> dict[str, Any]:
        for user in users:
            if user.get("sysUserId") is not None:
                return user
        return users[0]

    def _select_contact_way_primary_users(self, users: list[dict[str, Any]]) -> list[dict[str, Any]]:
        selected: list[dict[str, Any]] = []
        for expected_user_id in self._CONTACT_WAY_DEFAULT_USER_IDS:
            user = self._find_user_by_user_id(users, expected_user_id)
            if user is not None:
                selected.append(user)
        if selected:
            return selected
        if len(users) >= 2:
            return users[:2]
        if users:
            return [users[0]]
        raise RuntimeError("no_available_cp_users")

    def _select_contact_way_backup_user(
        self,
        users: list[dict[str, Any]],
        primary_users: list[dict[str, Any]],
    ) -> dict[str, Any]:
        preferred = self._find_user_by_user_id(users, self._CONTACT_WAY_DEFAULT_BACKUP_USER_ID)
        if preferred is not None:
            return preferred
        primary_ids = {str(user.get("userId")) for user in primary_users}
        for user in users:
            if str(user.get("userId")) not in primary_ids:
                return user
        if primary_users:
            return primary_users[0]
        raise RuntimeError("no_available_backup_user")

    def _find_user_by_user_id(self, users: list[dict[str, Any]], expected_user_id: str) -> dict[str, Any] | None:
        for user in users:
            if str(user.get("userId")) == expected_user_id:
                return user
        return None

    def _extract_department_id(self, user: dict[str, Any]) -> int | None:
        raw = user.get("departmentId")
        if isinstance(raw, int):
            return raw
        if not isinstance(raw, str):
            return None
        match = re.search(r"\d+", raw)
        return int(match.group(0)) if match else None

    def _normalize_int(self, value: object) -> int | None:
        if value is None:
            return None
        text = str(value).strip()
        if not text or text == "不知道":
            return None
        if text.isdigit():
            return int(text)
        return None


    def _runtime_login_headers(self, env: dict[str, str]) -> dict[str, str]:
        login_url = env.get("FRONT_LOGIN_PATH", "")
        if not login_url:
            raise RuntimeError("missing FRONT_LOGIN_PATH")
        login_body = env.get("FRONT_LOGIN_BODY_TEMPLATE", "")
        if not login_body:
            raise RuntimeError("missing FRONT_LOGIN_BODY_TEMPLATE")
        rendered_body = (
            login_body.replace("${username}", env.get("FRONT_USERNAME", ""))
            .replace("${password}", env.get("FRONT_PASSWORD", ""))
            .replace("${sysType}", env.get("FRONT_SYS_TYPE", ""))
        )
        payload = json.loads(rendered_body)
        headers = {"Content-Type": "application/json"}
        raw_headers = env.get("FRONT_LOGIN_HEADERS_JSON", "")
        if raw_headers and raw_headers not in {"4", "null", "None"}:
            headers.update({str(k): str(v) for k, v in json.loads(raw_headers).items()})
        response = self._request_json(
            env.get("FRONT_LOGIN_METHOD", "POST").upper(),
            login_url,
            payload=payload,
            headers=headers,
        )
        token_path = env.get("FRONT_TOKEN_JSON_PATH", "data.accessToken")
        token = self._extract_runtime_json_path(response, token_path)
        header_name = env.get("AUTH_HEADER_NAME", "Authorization")
        scheme = env.get("AUTH_SCHEME", "Bearer")
        if not scheme or scheme in {"无前缀", "NONE", "none"}:
            auth_value = str(token)
        else:
            auth_value = f"{scheme} {token}".strip()
        return {header_name: auth_value}

    def _extract_runtime_json_path(self, payload: dict[str, Any], path: str) -> object:
        current: object = payload
        for part in path.split("."):
            if not isinstance(current, dict) or part not in current:
                raise RuntimeError(f"missing json path: {path}")
            current = current[part]
        return current

    def _request_json(
        self,
        method: str,
        url: str,
        *,
        payload: dict[str, Any] | None = None,
        headers: dict[str, str] | None = None,
    ) -> dict[str, Any]:
        body = None if method.upper() == "GET" else json.dumps(payload or {}).encode("utf-8")
        request = urllib.request.Request(
            url=url,
            data=body,
            headers={"Content-Type": "application/json", **(headers or {})},
            method=method.upper(),
        )
        try:
            with urllib.request.urlopen(request, timeout=20) as response:
                raw = response.read().decode("utf-8")
        except urllib.error.HTTPError as exc:  # pragma: no cover - runtime dependent
            raw = exc.read().decode("utf-8")
            raise RuntimeError(f"http_error:{exc.code}:{raw}") from exc
        data = json.loads(raw)
        if not isinstance(data, dict):
            raise RuntimeError(f"invalid_json_response:{url}")
        if str(data.get("code")) not in {"00000", ""} and data.get("code") is not None:
            raise RuntimeError(f"business_error:{url}:{data.get('msg')}")
        return data

    def _build_pytest_env(self, template_config: dict[str, str] | None = None) -> dict[str, str]:
        env = {"PYTHONDONTWRITEBYTECODE": "1", **os.environ}
        profile = env.get("AI_PIPELINE_TEST_ENV_PROFILE")
        if profile:
            runtime = load_test_environment(profile, self.test_environment_config_path)
            if runtime:
                self._set_if_present(env, "API_BASE_URL", runtime.get("api_base_url"))
                auth = runtime.get("auth", {}) if isinstance(runtime.get("auth"), dict) else {}
                front = auth.get("front_user", {}) if isinstance(auth.get("front_user"), dict) else {}
                token = auth.get("token", {}) if isinstance(auth.get("token"), dict) else {}

                self._set_if_present(env, "FRONT_USERNAME", front.get("username"))
                self._set_if_present(env, "FRONT_PASSWORD", front.get("password"))
                self._set_if_present(env, "FRONT_LOGIN_PATH", front.get("login_path"))
                self._set_if_present(env, "FRONT_TOKEN_JSON_PATH", token.get("json_path"))
                self._set_if_present(env, "AUTH_HEADER_NAME", token.get("header_name"))
                self._set_if_present(env, "AUTH_SCHEME", token.get("scheme"))

        for key, value in (template_config or {}).items():
            if value:
                env[key] = str(value)
        return env

    def _set_if_present(self, env: dict[str, str], key: str, value: object) -> None:
        if value is not None and not env.get(key):
            env[key] = str(value)
